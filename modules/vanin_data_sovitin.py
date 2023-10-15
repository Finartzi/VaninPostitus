from openpyxl import load_workbook
from modules.operating_system import set_globals
from modules.excelohjaus import get_excel_workbooks, dict_structure, run_also_this
from modules.my_json_library import write_data_into_json_file

forenames = []
lastnames = []
addresses = []
postcodes = []
postoffices = []
email_addresses = []


def split_names(names):
    tmp = names.split()
    # tarkistetaan erikois-sukunimet
    checks = ['Von', 'von', 'Af', 'af', 'Van', 'van', 'Bin', 'bin', 'De', 'de', 'Al', 'al', 'Von der', 'von der', 'Van den', 'van den', 'Bin al', 'bin al']

    if tmp[0] in checks:
        lastname = tmp[0] + ' ' + tmp[1]
        forename = tmp[2]
    else:
        lastname = tmp[0]
        forename = tmp[1]

    return forename, lastname


def get_data(workbook_name, sheet_name):
    wb = load_workbook(workbook_name)
    ws = wb[sheet_name]
    values_list = list()

    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    for value in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
        values_list.append(value)

    return headers, values_list


def sort_out_data_to_json(headers, values, jsonfile):
    tmp = []
    headers.clear()
    dict_struct = dict_structure.copy()

    for value in values:
        forename, lastname = split_names(value[0])
        forenames.append(forename)
        lastnames.append(lastname)
        addresses.append(value[3])
        postcodes.append(value[4])
        postoffices.append(value[5])
        email_addresses.append(value[2])

    for i in range (0, len(addresses)):
        dict_struct["Etunimi:"] = forenames[i]
        dict_struct["Sukunimi:"] = lastnames[i]
        dict_struct["Osoite:"] = addresses[i]
        dict_struct["Postinumero:"] = postcodes[i]
        dict_struct["Postitoimipaikka:"] = postoffices[i]
        dict_struct["S\u00e4hk\u00f6postiosoite:"] = email_addresses[i]
        tmp.append(dict_struct)

    write_data_into_json_file(jsonfile, tmp)



def run_this():
    my_wb, my_new_wb, json_file = set_globals()
    # Hae "alkuperäiset" tiedot Excelistä
    sheet_names, _ = get_excel_workbooks(my_wb)
    headers, values = get_data(my_wb, sheet_names[1])
    sort_out_data_to_json(headers, values, json_file)

    run_also_this(my_new_wb, json_file)

    print(" --- ")
    print("Vanin data-sovitin valmis")