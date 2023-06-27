"""
    JSON-tiedostojen ohjaukseen liittyvä moduuli
    Koska JSON-tiedostojen ohjaukseen liittyy paljon mielenkiintoista tekniikkaa,
    katsoin parhaimmaksi laittaa nämä omaan moduliinsa.
    author: arto.paeivinen@gmail.com
    copyright: MIT
"""
import json
from openpyxl import load_workbook
from collections import OrderedDict


class CustomType:
    def __init__(self, key, value):
        self.key = key
        self.value = value
    
    # from sample this function as placeholder in class object
    def to_json(self):
        pass


def to_json(self):
    # Serialize the object custom object
    return json.dumps(self, default=lambda o: o.__dict__, sort_keys=False, indent=4)


def get_json_data(keys, data_slice):
    # Convert String array to json array
    result = []
    for item in data_slice:
        obj = CustomType(keys, item)
        result.append(json.loads(obj.to_json()))
    return json.dumps(result)


# Kopioidaan Excel-tiedosto JSON-tiedostoon
def dump_data_into_json_file(workbook_name, sheet_name, json_file):
    wb = load_workbook(workbook_name)
    ws = wb[sheet_name]
    data = []
    values_list = list()
    # print(ws.max_row)
    # print(ws.max_column)

    # otsikot
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    # print(headers)

    # tiedot
    for value in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
        values_list.append(value)

    # muodosta dictionary-objekti
    for row in values_list:
        data.append(OrderedDict(zip(headers, row)))

    write_data_into_json_file(json_file, data)


# haetaan tiedot JSON-tiedostosta (joka on kopio)
def get_data_from_json_file(json_file):
    values_list = []
    try:
        # avaa JSON/tiedosto
        with open(json_file) as json_file:
            data = json.load(json_file)

        # hae otsikot
        headers = list(data[0].keys())

        # hae tiedot
        for value in data:
            values_list.append(value.values())

        return headers, values_list
    except FileNotFoundError:
        print("JSON-tiedostoa ei löytynyt")
    except AttributeError:
        print("NoneType")


# Kirjoita Dictionary-objekti JSON-tiedostoon
def write_data_into_json_file(filename, data):
    # vie dictionary-objekti JSON-tiedostoon
    with open(filename, 'w') as outfile:
        json.dump(data, outfile, indent=4)
