"""
    Excel-taulukon ohjaukseen liittyvä moduuli
    author: arto.paeivinen@gmail.com
    copyright: MIT
"""
from modules import my_json_library
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Color, Fill, PatternFill
from modules.my_json_library import get_data_from_json_file, write_data_into_json_file

# tähän on laitettava polku käytettyyn tiedostoon, kuten se on koneella, jolla ohjelmaa suoritetaan
# my_wb = "C:\\Users\\artzi\\OneDrive\\Desktop\\Tarratesti.xlsx"  # työstettävän Excel-taulukon polku
my_wb = "/home/artzi/Desktop/Tarratesti.xlsx"                    # edellinen sama Linuxissa
my_new_wb = "/home/artzi/Desktop/Tarratesti2.xlsx"
json_file = "res/harjoitustiedot.json"                         # Uusi nimi tälle, kun oikeita tietoja käsitellään

keep_titels = False
keep_countries = False

json_email_file = "files/sähköpostitettavat.json"
json_mail_file = "files/kirjepostitettavat_kaikki.json"
json_final_mail_file = "files/kirjepostitettavat.json"
json_email_missing_file = "files/puuttuvat_spostiosoitteet.json"
json_phone_missing_file = "files/puuttuvat_puhelinnumerot.json"
json_problems_file = "files/tarkista_vaihtoehdot.json"
json_problems_file2 = "files/tarkista_käsin.json"

dict_structure = {
    "Titteli:": "",
    "Etunimi:": "",
    "Sukunimi:": "",
    "Osoite:": "",
    "Postinumero:": "",
    "Postitoimipaikka:": "",
    "Maa:": "",
    "S\u00e4hk\u00f6postitus:": "",
    "S\u00e4hk\u00f6postiosoite:": "",
    "Puhelinnumero:": ""}

label_structure = {
    "Nimi tai Nimet:": "",
    "Katuosoite:": "",
    "Postinumero ja -toimisto:": ""}

# Constants
SELECTION_EMAILS_ = 'Sähköpostitus:'
SELECTION_PHONES_ = 'Puhelinnumero:'
PARTLY_REMOVABLE_EMAILS_ = 'Sähköpostiosoite:'
COUNTRY_LITERAL_ = ' Finland'

titels = []
forenames = []
lastnames = []
addresses = []
postcodes = []
postoffices = []
countries = []
email_yes = []
email_addresses = []
phone_numbers = []
headers_values_main = []
posting_json_set = []
original_data_length = 0


# Yhdistetään Exceliin
def get_excel_workbooks(filename):
    workbook = load_workbook(filename)
    return workbook.sheetnames, workbook


# Putsataan dictionary-objektin tietoja hiukan
def cleanup_data(dict_value):
    x = str(dict_value)
    x = x.replace('dict_values', '')
    x = x.replace('([', '')
    x = x.replace('])', '')
    x = x.replace('"', '')
    x = x.replace("'", '')
    x = x.rstrip()
    x = x.lstrip()
    y = x.split(',')
    return y


# Putsataan string-objektin tietoja hiukan
def cleanup_string(value):
    x = value
    x = x.lstrip()
    x = x.replace('  ', ' ')
    return x


# Käsitellään tietokantaa ja kerätään tiedot koneen muistiin Arrayhin (nopeuttaa käsittelyä)
def set_values_into_arrays(file_name, only_posting=False):
    data = get_data_from_json_file(file_name)
    headers = list(map(str, data[0]))
    values_list = list(map(str, data[1]))

    for value in values_list:
        cleaned_row = cleanup_data(value)
        titels.append(cleaned_row[0])
        forenames.append(cleaned_row[1])
        lastnames.append(cleaned_row[2])
        addresses.append(cleaned_row[3])
        postcodes.append(cleaned_row[4])
        postoffices.append(cleaned_row[5])
        countries.append(cleaned_row[6])
        if not only_posting:
            email_yes.append(cleaned_row[7])
            email_addresses.append(cleaned_row[8])
            phone_numbers.append(cleaned_row[9])
    for item in headers:
        headers_values_main.append(item)
    return True


def clear_arrays():
    # reset arrays
    titels.clear()
    forenames.clear()
    lastnames.clear()
    addresses.clear()
    postcodes.clear()
    postoffices.clear()
    countries.clear()
    email_yes.clear()
    email_addresses.clear()
    phone_numbers.clear()
    headers_values_main.clear()


def amount_original_rows(json_set):
    return len(json_set)


# Luo uudet tiedostot sähköpostilistoille ja postituslistoille
def make_posting_json_files(json_set):

    def separate_missing_phones(json_type_set):
        phones = []
        for items in json_type_set:
            # Listaa puuttuvat puhelinnumerot
            if items[SELECTION_PHONES_] is None or items[SELECTION_PHONES_] == '':
                phones.append(items)
        return phones

    def separate_missing_emails(json_type_set):
        emails = []
        for items in json_type_set:
            # Listaa puuttuvat sähköpostiosoitteet
            if items[PARTLY_REMOVABLE_EMAILS_] is None or items[PARTLY_REMOVABLE_EMAILS_] == '':
                emails.append(items)
        return emails

    def enlist_emails(json_type_set, remove_selection_emails):
        emailing = []
        for items in json_type_set:
            # Listaa sähköpostitettavat
            if items[PARTLY_REMOVABLE_EMAILS_] is not None and \
                    items[PARTLY_REMOVABLE_EMAILS_] != "" and \
                    items[remove_selection_emails].lower() == "yes":
                # Listaa olemassa olevat sähköpostiosoitteet
                emailing.append(items)
        return emailing

    def enlist_posts(json_type_set, remove_selection_emails):
        posting = []
        problems = []
        for items in json_type_set:
            # Listaa tavan postitettavat
            if items[remove_selection_emails] is None or \
                    items[remove_selection_emails].lower() == "no" or \
                    items[remove_selection_emails] == "":
                posting.append(items)
            elif items[remove_selection_emails].lower() == "yes" and \
                    items[PARTLY_REMOVABLE_EMAILS_] is None or \
                    items[PARTLY_REMOVABLE_EMAILS_] == "":
                problems.append(items)
        return posting, problems

    missing_phones = separate_missing_phones(json_set)
    my_json_library.write_data_into_json_file(json_phone_missing_file, missing_phones)

    missing_emails = separate_missing_emails(json_set)
    my_json_library.write_data_into_json_file(json_email_missing_file, missing_emails)

    enlisted_emails = enlist_emails(json_set, SELECTION_EMAILS_)
    my_json_library.write_data_into_json_file(json_email_file, enlisted_emails)

    enlisted_post, manual_check = enlist_posts(json_set, SELECTION_EMAILS_)
    my_json_library.write_data_into_json_file(json_mail_file, enlisted_post)
    my_json_library.write_data_into_json_file(json_problems_file, manual_check)
    return True


def combine_key_value_pairs():
    slice_value = []
    key_value_pair = []

    def create_key_value_pairs(headers_array, slice_of_array_of_values):
        key_value_pairs = {}
        for ii in range(0, len(headers_array)):
            header = headers_array[ii]
            slices = slice_of_array_of_values.split(',')
            slices[ii] = slices[ii].lstrip()
            if slices[ii] == 'None':
                slices[ii] = None
                key_value_pairs[header] = slices[ii]
            else:
                key_value_pairs[header] = slices[ii]
        return key_value_pairs

    def create_slice_of_array_of_values(item_index):
        return f'{titels[item_index]}, ' \
               f'{forenames[item_index]}, ' \
               f'{lastnames[item_index]}, ' \
               f'{addresses[item_index]}, ' \
               f'{postcodes[item_index]}, ' \
               f'{postoffices[item_index]}, ' \
               f'{countries[item_index]}, ' \
               f'{email_yes[item_index]}, ' \
               f'{email_addresses[item_index]}, ' \
               f'{phone_numbers[item_index]}'

    for i in range(0, len(addresses)):
        slice_value.append(create_slice_of_array_of_values(i))
        key_value_pair.append(create_key_value_pairs(headers_values_main, slice_value[i]))
    return key_value_pair


# max. similar addresses is limited to 3 . More similar addresses must be checked manually
# Here we check how many problematic addresses we have
def sort_out_more_than_three(keep):
    # Here we find out which indexes they represent
    def find_out_more_than_three(keep_list, err_list):
        # in err_list we have starting points, now we look for the other relative points and
        # add start points to collection
        nums = set()
        for tuples in keep_list:
            if tuples[0] in err_list:
                nums.add(tuples[1])
            if tuples[1] in err_list:
                nums.add(tuples[0])
        for it in err_list:
            nums.add(it)
        return list(nums)

    counter = []
    sorter = set()
    for item in keep:
        counter.append(item[0])
    for num in counter:
        x = 0
        for n in counter:
            if n == num:
                x += 1
        if x > 2:
            sorter.add(num)
    return find_out_more_than_three(keep, sorter)


# find multiple posting addresses
def find_multiples():

    # make .json-file for working manually entries
    def create_file_to_manually_workout(inx_list, file_name):
        set_values_into_arrays(json_file)
        tmp = []
        for i in range(0, len(addresses)-1):
            tmp_dict = dict_structure.copy()
            if i in inx_list:
                tmp_dict["Titteli:"] = titels[i]
                tmp_dict["Etunimi:"] = forenames[i]
                tmp_dict["Sukunimi:"] = lastnames[i]
                tmp_dict["Osoite:"] = addresses[i]
                tmp_dict["Postinumero:"] = postcodes[i]
                tmp_dict["Postitoimipaikka:"] = postoffices[i]
                tmp_dict["Maa:"] = countries[i]
                tmp_dict["S\u00e4hk\u00f6postitus:"] = email_yes[i]
                tmp_dict["S\u00e4hk\u00f6postiosoite:"] = email_addresses[i]
                tmp_dict["Puhelinnumero:"] = phone_numbers[i]
                tmp.append(tmp_dict)
        write_data_into_json_file(file_name, tmp)

    # here we mark tuples to be removed according given indexes
    def find_tuples_to_remove(inx_list, keep_list, double):
        safe = []
        safe_list = []
        for ix in range(0, len(keep_list)):
            x = keep_list[ix]
            if (x[0] or x[1]) not in inx_list:
                safe.append(keep_list[ix])
        for iy in double:
            if iy not in inx_list:
                safe_list.append(iy)
        return safe_list, safe

    def find_doubles_and_more():
        keeper_tuples = []
        doubl_list = set()
        # find doubles and maybe triples
        for i in range(0, len(addresses)):
            j = 0
            for j in range(j, len(addresses)):
                if i != j:
                    if postcodes[i] == postcodes[j] and addresses[i] == addresses[j]:
                        doubl_list.add(i)
                        doubl_list.add(j)
                        if i < j:
                            keeper_tuples.append([i, j])
        return doubl_list, keeper_tuples

    doubles, keeper = find_doubles_and_more()
    problem_indexes = sort_out_more_than_three(keeper)
    create_file_to_manually_workout(problem_indexes, json_problems_file2)
    doubles_list, doubles_tuples = find_tuples_to_remove(problem_indexes, keeper, list(doubles))

    return doubles_list, doubles_tuples, problem_indexes


# find single posting addresses
def find_singles(arr, multiples, issues, file_name):
    clear_arrays()
    set_values_into_arrays(file_name)
    single = set()
    for i in range(0, len(arr)):
        if i not in issues and i not in multiples:
            single.add(i)
    return list(sorted(single))


# find up to three similar addresses
def find_triplets(doubles):
    triplet = []
    aa = set()
    bb = []
    for i in range(0, len(doubles)):
        tmp = doubles[i][0]
        # looking for 3 similar addresses
        for j in range(0, len(doubles)):
            if i != j and doubles[j][0] == tmp:
                aa.add(doubles[i][0])
                aa.add(doubles[i][1])
                bb.append([doubles[i]])
    a = 0
    b = 0
    for items in bb:
        for item in items:
            if item[0] != a:
                a = item[0]
                b = item[1]
            else:
                c = item[1]
                triplet.append(sorted({a, b, c}))
    return triplet


# find the doubles excluding triples
def find_double(tripple, keeper):
    exclude = set()
    exclude_indexes = set()
    output = []
    # find only doubles
    for item in tripple:
        for e in item:
            exclude.add(e)
        for i in range(0, len(keeper)):
            if keeper[i][0] in exclude:
                exclude_indexes.add(i)
    for i in range(0, len(keeper)):
        if i not in exclude_indexes:
            output.append(keeper[i])
    return output


# get single info for labels
def single_labels(index_values, title=False, country=False):
    labels = []
    for i in range(0, len(index_values)):
        label = []
        if title and (titels[index_values[i]] != 'None'):
            combined_name = f'{titels[index_values[i]]} {forenames[index_values[i]]} {lastnames[index_values[i]]}'
        else:
            combined_name = f'{forenames[index_values[i]]} {lastnames[index_values[i]]}'
        label.append(cleanup_string(combined_name))
        label.append(cleanup_string(addresses[index_values[i]]))
        combined_postal = f'{postcodes[index_values[i]]} {postoffices[index_values[i]]}'
        label.append(cleanup_string(combined_postal))
        if countries[index_values[i]] != COUNTRY_LITERAL_ or country:
            label.append(cleanup_string(countries[index_values[i]]))
        labels.append(label)
    return labels


# combine double info's for templates
def double_labels(index_values, title=False, country=False):
    labels = []
    for i in range(0, len(index_values)):
        label = []
        if titels[index_values[i][0]] != 'None':
            title1 = titels[index_values[i][0]]
        else:
            title1 = ''
        if titels[index_values[i][1]] != 'None':
            title2 = titels[index_values[i][1]]
        else:
            title2 = ''
        if title:
            if lastnames[index_values[i][0]] == lastnames[index_values[i][1]]:
                combined_name = f'{title1} {forenames[index_values[i][0]]} & ' \
                                f'{title2} {forenames[index_values[i][1]]} {lastnames[index_values[i][0]]}'
            else:
                combined_name = f'{title1} {forenames[index_values[i][0]]} {lastnames[index_values[i][0]]} & ' \
                                f'{title2} {forenames[index_values[i][1]]} {lastnames[index_values[i][1]]}'
        else:
            if lastnames[index_values[i][0]] == lastnames[index_values[i][1]]:
                combined_name = f'{forenames[index_values[i][0]]} & ' \
                                f'{forenames[index_values[i][1]]} {lastnames[index_values[i][0]]}'
            else:
                combined_name = f'{forenames[index_values[i][0]]} {lastnames[index_values[i][0]]} & ' \
                                f'{forenames[index_values[i][1]]} {lastnames[index_values[i][1]]}'
        label.append(cleanup_string(combined_name))
        label.append(cleanup_string(addresses[index_values[i][0]]))
        combined_postal = f'{postcodes[index_values[i][0]]} {postoffices[index_values[i][0]]}'
        label.append(cleanup_string(combined_postal))
        if countries[index_values[i][0]] != COUNTRY_LITERAL_ or country:
            label.append(cleanup_string(countries[index_values[i][0]]))
        labels.append(label)
    return labels


# combine triple-info's for templates no titles here cause no space on labels
def triple_labels(index_values, country=False):
    labels = []
    for i in range(0, len(index_values)):
        label = []
        combined_name = ''
        # here we use no titles at all
        if lastnames[index_values[i][0]] == lastnames[index_values[i][1]] == lastnames[index_values[i][2]]:
            combined_name = f'{forenames[index_values[i][0]]} & ' \
                            f'{forenames[index_values[i][1]]} & ' \
                            f'{forenames[index_values[i][2]]} {lastnames[index_values[i][0]]}'
        elif lastnames[index_values[i][0]] != lastnames[index_values[i][1]] != lastnames[index_values[i][2]]:
            combined_name = f'{forenames[index_values[i][0]]} {lastnames[index_values[i][0]]} & ' \
                            f'{forenames[index_values[i][1]]} {lastnames[index_values[i][1]]} & ' \
                            f'{forenames[index_values[i][2]]} {lastnames[index_values[i][2]]}'
        elif lastnames[index_values[i][0]] == lastnames[index_values[i][1]] != lastnames[index_values[i][2]]:
            combined_name = f'{forenames[index_values[i][0]]} & ' \
                            f'{forenames[index_values[i][1]]} {lastnames[index_values[i][0]]} & ' \
                            f'{forenames[index_values[i][2]]} {lastnames[index_values[i][2]]} '
        elif lastnames[index_values[i][0]] != lastnames[index_values[i][1]] == lastnames[index_values[i][2]]:
            combined_name = f'{forenames[index_values[i][0]]} {lastnames[index_values[i][0]]} & ' \
                            f'{forenames[index_values[i][1]]} & ' \
                            f'{forenames[index_values[i][2]]} {lastnames[index_values[i][1]]}'
        else:
            print('This should not happen')
        label.append(cleanup_string(combined_name))
        label.append(cleanup_string(addresses[index_values[i][0]]))
        combined_postal = f'{postcodes[index_values[i][0]]} {postoffices[index_values[i][0]]}'
        label.append(cleanup_string(combined_postal))
        if countries[index_values[i][0]] != COUNTRY_LITERAL_ or country:
            label.append(cleanup_string(countries[index_values[i][0]]))
        labels.append(label)
    return labels


def all_labels(singles, doubles, triples):
    _all = []
    for item in singles:
        _all.append(item)
    for item in doubles:
        _all.append(item)
    for item in triples:
        _all.append(item)
    return _all


# Here is where the magic happens!
def sort_postal_addresses(file_name):
    # clear old Arrays
    clear_arrays()
    # get posting data-Arrays
    set_values_into_arrays(file_name, True)
    # holder keeps two values pointing Array-indexes containing similar address
    doubles_and_triples, holder, problem_indexes = find_multiples()
    # find single posting addresses
    # problem_indexes = sort_out_more_than_three(doubles_and_triples)
    singles = find_singles(addresses, doubles_and_triples, problem_indexes, file_name)
    # find triple posting addresses
    triplets = find_triplets(holder)
    # find double posting addresses
    doubles = find_double(triplets, holder)
    single_labeled = single_labels(singles, keep_titels, keep_countries)
    double_labeled = double_labels(doubles, keep_titels, keep_countries)
    triple_labeled = triple_labels(triplets, keep_countries)
    all_labeled = all_labels(single_labeled, double_labeled, triple_labeled)
    return all_labeled


def info(ready_labels, original_set):
    a = len(ready_labels)
    b = amount_original_rows(original_set)
    c = 100 - (a * 100 / b)
    print('--- ')
    print(f'Tulostettavia etikettejä: {a}')
    print(f'Alkuperäisien tietojen määrä: {b}')
    print(f'Tulostettavia etikettejä säästynyt: {b - a}')
    print(f'Säästö prosentteina: {c:.2f} %')


def write_excel_file(excel_file_name):
    email = "Laita_sähköpostia"
    mail = "Laita_kirjepostia"
    missing_email = "Puuttuu_sähköpostiosoite"
    missing_phone = "Puuttuu_puhelinnumero"
    check_alternatives = "Tarkista_vaihtoehdot"
    check_manually = "Tarkista_osoitteet"

    workbook = Workbook()
    workbook = write_excel_sheet(workbook, email, json_email_file)
    workbook = write_excel_sheet(workbook, mail, json_final_mail_file)
    workbook = write_excel_sheet(workbook, missing_email, json_email_missing_file)
    workbook = write_excel_sheet(workbook, missing_phone, json_phone_missing_file)
    workbook = write_excel_sheet(workbook, check_alternatives, json_problems_file)
    workbook = write_excel_sheet(workbook, check_manually, json_problems_file2)

    workbook.save(excel_file_name)
    workbook.close()


# Kirjoitetaan dataa excelin taulukoihin
def write_excel_sheet(workbook_object, excel_sheet_name, json_file):
    json_data = []
    headers, values_list = get_data_from_json_file(json_file)

    header_style_settings = PatternFill('solid', fgColor='FFFF00')
    header_font_settings = Font(bold=True, size=16, color='0000FF')
    data_font_settings = Font(bold=False, size=10, color='808080')

    # if not excel_sheet_name in sheet_names:
    ws = workbook_object.create_sheet(excel_sheet_name)
    # else:
    #     ws = workbook.get_sheet_by_name(excel_sheet_name)
    for value in values_list:
        cleaned_row = cleanup_data(value)
        json_data_set = []
        for item in cleaned_row:
            json_string = cleanup_string(item)
            json_data_set.append(json_string)
        json_data.append(json_data_set)

    for i in range(0, len(headers)):
        col = i + 1
        fix = 1
        ws.cell(row=fix, column=col, value=headers[i]).font = header_font_settings
        ws.cell(row=fix, column=col).fill = fill = PatternFill('solid', bgColor='FFFF00')
        for j in range(0, len(values_list)):
            r = j + 2
            ws.cell(row=r, column=col, value=json_data[j][i]).font = data_font_settings
    return workbook_object


def create_modified_headers_for_labels(data):
    obj = []
    for i in range(0, len(data)):
        dic = label_structure.copy()
        dic["Nimi tai Nimet:"] = data[i][0]
        dic["Katuosoite:"] = data[i][1]
        dic["Postinumero ja -toimisto:"] = data[i][2]
        obj.append(dic)
    return obj


def run_this():
    # # Seuraavien 2 rivin aktivointi hakee "alkuperäiset" tiedot Excelistä
    # sheet_names, wb = get_excel_workbooks(my_wb)
    # my_json_library.dump_data_into_json_file(my_wb, sheet_names[0], json_file)
    #
    # Varmuuskopioidusta Excelistä, joka on nyt muutettu JSON-tiedostoksi, haetaan tietoja koneluettavaksi
    my_json_library.get_data_from_json_file(json_file)
    set_values_into_arrays(json_file)
    # Luodaan JSON-tyyppistä dataa koneen muistiin
    valid_json_set = combine_key_value_pairs()
    # Luodaan sekä sähköposteille että postituslistalle omat JSON-tiedostot (myös vikailmoitukset)
    make_posting_json_files(valid_json_set)
    # Lajitellaan postitukset
    labels = sort_postal_addresses(json_mail_file)
    info(labels, valid_json_set)
    # vie valmiit etikettitiedot JSON-tiedostoon
    my_json_library.write_data_into_json_file(json_final_mail_file, create_modified_headers_for_labels(labels))
    # luodaan uusia excel-taulukoita
    write_excel_file(my_new_wb)

    print(" --- ")
    print("Valmis")
